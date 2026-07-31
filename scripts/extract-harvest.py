#!/usr/bin/env python3
"""Extract harvest workbook -> scripts/harvest-data.json for import-harvest.js."""
import json, sys, datetime
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else '/opt/data/cache/documents/doc_9e7be28281a9_GGH Members Only Harvest.xlsx'
OUT = '/opt/data/TreeBot/scripts/harvest-data.json'

wb = openpyxl.load_workbook(XLSX, data_only=True)

# --- SEEDLOTS: species rows x areas 1-26 ---
ws = wb['SEEDLOTS']
areas = [int(c.value) for c in ws[2][2:28]]  # row 2: area numbers 1..26
seedlots_species = []
allocations = []
for row in ws.iter_rows(min_row=3, max_row=28, values_only=True):
    species = row[0]
    if not species:
        continue
    seedlots_species.append(species.strip())
    for i, area in enumerate(areas):
        v = row[2 + i]
        if isinstance(v, (int, float)) and v > 0:
            allocations.append({'species': species.strip(), 'area': area, 'stems': int(v)})

# --- TREE TRACKER: drops ---
ws = wb['TREE TRACKER']
drops = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    date, area, crew, seedlot, boxes, partial1, ret, partial2 = row[:8]
    if not seedlot or not area:
        continue
    seedlot = str(seedlot).strip().lower()
    # partial stems: column F (# in Partial Box) accompanies taken boxes.
    # column H (# in Partial Box after returned) is rare; row 251 special-cases
    # boxes=None partial=10 meaning 10 loose stems returned *out* — treat as a
    # drop of 0 boxes + 10 partial stems taken (the sheet's own total col shows -10).
    partial_stems = 0
    if isinstance(partial1, (int, float)):
        partial_stems = int(partial1)
    b = int(boxes) if isinstance(boxes, (int, float)) else 0
    if b == 0 and partial_stems == 0:
        continue
    drops.append({
        'rowNum': i,
        'date': date.strftime('%Y-%m-%d %H:%M:%S') if isinstance(date, datetime.datetime) else '2026-07-26 00:00:00',
        'area': int(area),
        'crew': str(crew).strip() if crew else '?',
        'seedlot': seedlot,
        'boxes': b,
        'partialStems': partial_stems,
    })

# --- THE PLAN: reefer stock ---
ws = wb['THE PLAN']
reefer = []
for row in ws.iter_rows(min_row=3, max_row=23, values_only=True):
    # cols: Area Code | Species | Total stems | Total Boxes | Boxes at Reefer | Partials
    species = row[1]
    if not species:
        continue
    at_reefer = row[4]
    rb = at_reefer if isinstance(at_reefer, (int, float)) else 0
    reefer.append({'species': str(species).strip(), 'reeferBoxes': int(rb)})

out = {
    'seedlotsSpecies': seedlots_species,
    'allocations': allocations,
    'drops': drops,
    'reefer': reefer,
}
with open(OUT, 'w') as f:
    json.dump(out, f, indent=1)
print(f"species={len(seedlots_species)} allocations={len(allocations)} drops={len(drops)} reefer={len(reefer)}")
print("stems total:", sum(a['stems'] for a in allocations))
print("drop boxes:", sum(d['boxes'] for d in drops), "partials:", sum(d['partialStems'] for d in drops))
